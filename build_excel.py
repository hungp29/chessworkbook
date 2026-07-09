from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.page import PageMargins

ROOT_DIR = Path("./img_root")

# ==================================================
# CONFIG
# ==================================================

IMAGE_WIDTH = 130
IMAGE_HEIGHT = 130

# 3 ảnh / hàng
IMAGE_COLS = ["A", "D", "G"]

# ==================================================
# WORKBOOK
# ==================================================

wb = Workbook()
ws = wb.active
ws.title = "Chess Workbook"

# ==================================================
# A4 SETTINGS
# ==================================================

ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.orientation = "portrait"

ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0

ws.page_margins = PageMargins(
    left=0.2,
    right=0.2,
    top=0.3,
    bottom=0.3,
)

# ==================================================
# COLUMN WIDTHS
# ==================================================

for col in ["A", "B", "C"]:
    ws.column_dimensions[col].width = 9

for col in ["D", "E", "F"]:
    ws.column_dimensions[col].width = 9

for col in ["G", "H", "I"]:
    ws.column_dimensions[col].width = 9

# ==================================================
# TABLE OF CONTENTS
# ==================================================

row = 1

ws.merge_cells("A1:I1")

cell = ws["A1"]
cell.value = "TABLE OF CONTENTS"
cell.font = Font(size=20, bold=True)

row += 2

chapters = sorted(
    p for p in ROOT_DIR.iterdir()
    if p.is_dir()
)

for chapter in chapters:

    group_file = chapter / "group.txt"

    chapter_name = (
        group_file.read_text(encoding="utf-8").strip()
        if group_file.exists()
        else chapter.name
    )

    ws.merge_cells(
        start_row=row,
        start_column=1,
        end_row=row,
        end_column=9
    )

    cell = ws.cell(row=row, column=1)
    cell.value = chapter_name
    cell.font = Font(bold=True)

    row += 1

    items = sorted(
        p for p in chapter.iterdir()
        if p.is_dir() and p.name.startswith("item")
    )

    for index, item in enumerate(items, start=1):

        title_file = item / "title.txt"

        title = (
            title_file.read_text(encoding="utf-8").strip()
            if title_file.exists()
            else item.name
        )

        ws.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=9
        )

        ws.cell(
            row=row,
            column=1,
            value=f"    {index}. {title}"
        )

        row += 1

row += 3

# ==================================================
# CONTENT
# ==================================================

for chapter in chapters:

    group_file = chapter / "group.txt"

    chapter_name = (
        group_file.read_text(encoding="utf-8").strip()
        if group_file.exists()
        else chapter.name
    )

    # ------------------------------------------
    # CHAPTER TITLE
    # ------------------------------------------

    ws.merge_cells(
        start_row=row,
        start_column=1,
        end_row=row,
        end_column=9
    )

    cell = ws.cell(row=row, column=1)

    cell.value = chapter_name
    cell.font = Font(
        size=18,
        bold=True
    )

    row += 2

    items = sorted(
        p for p in chapter.iterdir()
        if p.is_dir() and p.name.startswith("item")
    )

    for index, item in enumerate(items, start=1):

        title_file = item / "title.txt"

        title = (
            title_file.read_text(encoding="utf-8").strip()
            if title_file.exists()
            else item.name
        )

        # ------------------------------------------
        # ITEM TITLE
        # ------------------------------------------

        ws.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=9
        )

        title_cell = ws.cell(
            row=row,
            column=1
        )

        title_cell.value = f"{index}. {title}"

        title_cell.alignment = Alignment(
            wrap_text=True,
            vertical="top"
        )

        title_cell.font = Font(
            size=11,
            bold=False
        )

        row += 2

        # ------------------------------------------
        # IMAGES
        # ------------------------------------------

        images = sorted(
            (item / "cropped").glob("*.jpg")
        )

        rows_needed = (
            len(images) + 2
        ) // 3

        for img_index, img_file in enumerate(images):

            excel_col = IMAGE_COLS[
                img_index % 3
            ]

            excel_row = (
                row
                + (img_index // 3) * 10
            )

            img = XLImage(str(img_file))

            img.width = IMAGE_WIDTH
            img.height = IMAGE_HEIGHT

            ws.add_image(
                img,
                f"{excel_col}{excel_row}"
            )

        row += rows_needed * 10

        row += 2

    row += 3

# ==================================================
# PRINT AREA
# ==================================================

ws.print_area = f"A1:I{row}"

# ==================================================
# SAVE
# ==================================================

wb.save("Chess Workbook.xlsx")

print("Created Chess Workbook.xlsx")